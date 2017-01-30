from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color
from openpyxl.styles import colors, PatternFill, Alignment, Border, Side
import sys, string
wb = Workbook()
ws1 = wb.active
ws1.title = "Week 1"

#Create lists for the outter border to run through and add a border
outline_right = [1,2, 4, 5, 6, 33,34, 35]
outline_left = [1,2,4,6, 32, 33,35]
outline_top = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
outline_bottom = ['B', 'C', 'D', 'E', 'F', 'G', 'H']

#------------------------------------------------

#Function adds bottom border or bottom and side border
def cellBot(row, colStart, colEnd, side):
    abc = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    r = row
    cs = abc.index(colStart)
    ce = abc.index(colEnd)
    ce += 1

    for num in range(cs, ce):
        col = abc[num]
        cell = col + r
        cell = '%s' %cell

        if side == 'R':
            ws1[cell].border = Border(bottom = Side(style = 'thin'),
                                      right = Side(style = 'thick'))
        elif side == 'L':
            ws1[cell].border = Border(bottom = Side(style= 'thin'),
                                      left = Side(style = 'thick'))
        else:
            ws1[cell].border = thin_bottom
    
#--------------------------------------------------

#Function adds full borders to table
def cellBox(startR, endR, startC, endC):
    abc = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    
    sRow = startR
    eRow = endR
    sCol = startC   
    eCol = endC
    eRow += 1
    cs = abc.index(sCol)
    ce = abc.index(eCol)
    ce += 1
    x = sRow

    while x <= eRow:
        for i in range(cs, ce):
            col = abc[i]
            cell = col + str(x)
            cell = '%s' %cell
            if cell == "G7" or cell == "H7":
                ws1[cell].border = no_bottom
            elif cell == "G8" or cell == "H8":
                ws1[cell].border = no_top
            else:
                ws1[cell].border = thin_border
        x += 1
    

#------------------------------------------------------

#Saves Excel sheet under name
dest_filename = 'Hourlog_test.xlsx'

#---------------------------------------------------------

#States the size of the borders
thin_bottom = Border(bottom=Side(style='thin'))
thin_border = Border(left = Side(style = 'thin'),
                    right=Side(style = 'thin'),
                    top=Side(style = 'thin'),
                    bottom=Side(style = 'thin'))
no_bottom = Border(left = Side(style = 'thin'),
                    right=Side(style = 'thin'),
                    top=Side(style = 'thin'))
no_top = Border(left = Side(style = 'thin'),
                    right=Side(style = 'thin'),
                    bottom=Side(style = 'thin'))
thin_left = Border(left = Side(style = 'thin'))
thin_right = Border(right = Side(style = 'thin'))
thin_top = Border(top = Side(style = 'thin'))

thick_top = Border(top = Side(style = 'thick'))
thick_bottom = Border(bottom = Side(style = 'thick'))
thick_left = Border(left = Side(style = 'thick'))
thick_right = Border(right = Side(style = 'thick'))
                   

#--------------------------------------------------------

#Merges the Table Labels together to allow labels to be stacked and centered
for i in range(0, 9):
    alph = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    col = alph[i]
    cell = ws1[col + '7']
    cell2 = ws1[col + '8']
    if col == "G" or col == "H":
        cell.alignment = Alignment(horizontal ='center', vertical = 'center',
                                   text_rotation= 0, wrap_text = False,
                                   shrink_to_fit = False, indent = 0)
        cell2.alignment = Alignment(horizontal = 'center', vertical = 'center',
                                    text_rotation= 0, wrap_text = False,
                                    shrink_to_fit=False, indent = 0)
        continue
    ws1.merge_cells(col + '7:' + col + '8')
    cell.alignment = Alignment(horizontal = 'center', vertical = 'center',
                               text_rotation = 0, wrap_text = False,
                               shrink_to_fit = False, indent = 0)

#--------------------------------------------------------
    
#calls the borders function to add borders to the whole table
cellBox(7, 29, 'A', 'I')


#--------------------------------------------------------

#Adjusts the sizes of the cells
ws1.column_dimensions['E'].width = 19
ws1.column_dimensions['A'].width = 9
ws1.column_dimensions['C'].width = 9
ws1.column_dimensions['D'].width = 10
ws1.column_dimensions['F'].width = 17
ws1.column_dimensions['G'].width = 25
ws1.column_dimensions['H'].width = 25
ws1.column_dimensions['I'].width = 40

#-------------------------------------------------------

#calls the merge function for SHEET LABEL
ws1.merge_cells('A1:I1')

#------------------------------------------------------
a1 = ws1['A1']
a1.fill = PatternFill("solid", fgColor = "2F75B5")
ws1['A1'] = "Electrical Trainee Bi-Weekly Work Log"
a1.font = Font(size = 16, color = "FFFFFF", bold = True)
a1.alignment = Alignment(horizontal = "center", vertical ='center', text_rotation= 0,
             wrap_text=False, shrink_to_fit =False, indent =0)

ws1['A3'] = "Trainee Name: "
ws1.merge_cells('A3:D3')
a3 = ws1['A3']
a3.font = Font(bold = True)
cellBot('3', 'A', 'D', 'L')

ws1['F3'] = "Payroll Start Date: "
ws1['F3'].font = Font(bold = True)
ws1.merge_cells('F3:G3')
cellBot('3', 'F', 'G', 'N')

ws1['I3'] = "Payroll End Date: "
ws1['I3'].font = Font(bold = True)
cellBot('3', 'I', 'I', 'R')

ws1['A5'] = "License Number: "
ws1['A5'].font = Font(bold = True)
ws1.merge_cells('A5:D5')
cellBot('5', 'A', 'D', 'L')
ws1['F5'] = "License Type: "
ws1['F5'].font = Font(bold = True)
ws1.merge_cells('F5:G5')
cellBot('5', 'F', 'G', 'N')

ws1['A7'] = "Date"
ws1['A7'].font = Font(size = 10, bold = True)
ws1['B7'] = "Start Time"
ws1['B7'].font = Font(size = 10, bold = True)
ws1['C7'] = "End Time"
ws1['C7'].font = Font(size = 10, bold = True)
ws1['D7'] = "Total Hours"
ws1['D7'].font = Font(size = 10, bold = True)
ws1['E7'] = "Work Order Number"
ws1['E7'].font = Font(size = 10, bold = True)
ws1['F7'] = "(01hr)Permit#/AHJ"
ws1['F7'].font = Font(size = 10, bold = True)
ws1['G7'] = "Journeyman Name"
ws1['G7'].font = Font(size = 10, bold = True)
ws1['G8'] = "(Printed)"
ws1['G8'].font = Font(size = 10, bold = True)
ws1['H7'] = "Journeyman"
ws1['H7'].font = Font(size = 10, bold = True)
ws1['H8'] = "Signature"
ws1['H8'].font = Font(size = 10, bold = True)
ws1['I7'] = "Work Description"
ws1['I7'].font = Font(size = 10, bold = True)

ws1['A31'] = "Sheet Summery"
ws1['A31'].font = Font(bold = True)
ws1.merge_cells('A31:C31')
cellBox(30, 30, 'A', 'D')
ws1['A31'].alignment = Alignment(horizontal = "center", vertical ='center')

ws1['F31'] = "Journeyman(1) Lic. #"
ws1['F31'].font = Font(bold = True)
ws1.merge_cells('F31:G31')
cellBot('31', 'F', 'G','N')
ws1['I31'] = "Journeyman(3) Lic. #"
ws1['I31'].font = Font(bold = True)
cellBot('31', 'I', 'I', 'R')
                
ws1['F32'] = "Journeyman(2) Lic. #"
ws1['F32'].font = Font(bold = True)
ws1.merge_cells('F32:G32')
cellBot('32', 'F', 'G', 'N')
ws1['I32'] = "Journeyman(4) Lic. #"
ws1['I32'].font = Font(bold = True)
cellBot('32', 'I', 'I','R')


ws1['A34'] = "Trainee Signature: "
ws1['A34'].font = Font(bold = True)
ws1.merge_cells('A34:D34')
cellBot('34', 'A', 'D', 'L')

ws1['F34'] = "Date: "
ws1['F34'].font = Font(bold = True)
cellBot('34', 'F', 'G','N')

ws1['I35'] = "CBRE"
i35 = ws1['I35']
i35.font = Font(size=24, color = "2F75B5", bold = True)

#--------------------------------------------------------
#################### NEED HELP #######################################
#adds border to the whole sheet
for j in outline_right:
    num = str(j)
    i = 'I' + num
    right = ws1[i]
    if right == 'I1':
        ws1[right].border = Border(top = Side(style = 'thick'),
                                  right = Side(style = 'thick'))
    elif right == 'I35':
        ws1[right].border = Border(right = Side(style = 'thick'),
                                   bottom = Side(style = 'thick'))
    else:
        right.border = thick_right
    
for i in outline_left:
    num = str(i)
    a = 'A' + num
    left = ws1[a]

    if left == 'A1':
        ws1[left].border = Border(top = Side(style = 'thick'),
                                  left = Side(style = 'thick'))
    elif left == 'A35':
        ws1[left].border = Border(bottom = Side(style = 'thick'),
                                   left =Side(style = 'thick'))
    else:
        left.border = thick_left

for k in outline_top:
    b = k + '1'
    top = ws1[b]
    top.border = thick_top

for y in outline_bottom:
    c = y + '35'
    bottom = ws1[c]
    bottom.border = thick_bottom
    
#--------------------------------------------------------
wb.save(filename = dest_filename)
ws1.print_area = 'A1:I35'
